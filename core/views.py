
# core/views.py
# ---------------------------------------------
# Django views for tfar1 (multi-tenant, Railway-ready)
# Upload .xlsx (15 columns) OR 16 including 'client', validate, store, display, download CSV
# ---------------------------------------------

import io
from typing import Any, List

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404

from .forms import UploadForm, ClientSelectForm
from .models import TfarRecord, Client, ClientMembership

# Base required headers (case-insensitive, any order allowed in this Option B impl)
REQUIRED_HEADERS: List[str] = [
    "asset id",
    "asset description",
    "tax start date",
    "depreciation method",
    "purchase cost",
    "tax effective life",
    "opening cost",
    "opening accumulated depreciation",
    "opening wdv",
    "addition",
    "disposal",
    "tax depreciation",
    "closing cost",
    "closing accumulated depreciation",
    "closing wdv",
]
OPTIONAL_CLIENT_HEADER = "client"


def _cell_value(row: tuple[Cell | Any, ...], index: int) -> Any:
    """Safe accessor for an Excel row tuple."""
    try:
        return row[index]
    except Exception:
        return None


def _to_int(value: Any) -> int:
    """Convert Excel cell value to int, defaulting to 0 for blanks / None."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return 0
    try:
        # Excel numeric cells can be float; cast safely
        return int(round(float(value)))
    except Exception:
        raise ValueError(f"Cannot convert '{value}' to integer")


def _to_str(value: Any, max_len: int) -> str:
    """Convert to string (safe), truncate to max_len."""
    s = "" if value is None else str(value)
    return s[:max_len]


def _to_date(value: Any):
    """
    Convert Excel cell value to Python date.
    openpyxl returns datetime.date or datetime.datetime for date cells automatically when data_only=True.
    If it's a string, try ISO parse.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise ValueError("Tax Start Date is required")

    # If already a date/datetime-like object
    try:
        # attempt to access .date() (datetime)
        if hasattr(value, "date"):
            return value.date()
    except Exception:
        pass

    # If it's already a date (openpyxl may use date class)
    from datetime import date
    if isinstance(value, date):
        return value

    # Fallback: attempt string parse (ISO-like)
    from datetime import datetime
    try:
        return datetime.fromisoformat(str(value)).date()
    except Exception:
        raise ValueError(f"Cannot parse date '{value}'")


def login_view(request):
    if request.method == "POST":
        u = request.POST.get("username"); p = request.POST.get("password")
        user = authenticate(request, username=u, password=p)
        if user: login(request, user); return redirect("dashboard")
        return render(request, "login.html", {"error": "Wrong login or password. Don't try again..."})
    return render(request, "login.html")

def logout_view(request):
    logout(request); return redirect("login")

###TEMP - DELETE FOR PROD
import traceback

def safe_debug(request):
    try:
        # Force dashboard code
        rows = TfarRecord.objects.filter(owner=request.user)[:1]
        return HttpResponse("Dashboard query OK, count=" + str(rows.count()))
    except Exception as e:
        return HttpResponse(
            "<pre>" + str(e) + "\n\n" + traceback.format_exc() + "</pre>", 
            content_type="text/plain"
        )

def debug_view(request):
    raise Exception("Manual test exception — this should show a 500 if debug middleware isn’t interfering")

####TEMP ABOVE




@login_required
def dashboard(request):
    memberships = (
        ClientMembership.objects.filter(user=request.user)
        .select_related("client")
        .order_by("client__name")
    )

    # Guard: no client memberships
    if not memberships.exists():
        return render(
            request,
            "dashboard.html",
            {
                "rows": [],
                "form": None,
                "client": None,
                "error": "You are not assigned to any clients. Please ask an administrator to add you.",
            },
        )

    # Determine selected client (from POST or session), fallback to first membership
    selected_client_id = request.POST.get("client") or request.session.get("selected_client_id")
    if not selected_client_id:
        selected_client_id = str(memberships.first().client.id)

    # Persist selection in session
    request.session["selected_client_id"] = selected_client_id

    client = get_object_or_404(Client, id=selected_client_id)

    # Guard: ensure user belongs to this client
    if not memberships.filter(client=client).exists():
        return render(
            request,
            "dashboard.html",
            {"rows": [], "form": ClientSelectForm(user=request.user), "error": "You don't have access to this client."},
        )

    rows = (
        TfarRecord.objects.filter(owner=request.user, client=client)
        .order_by("-uploaded_at", "asset_id")[:500]
    )

    form = ClientSelectForm(user=request.user, data={"client": selected_client_id})
    return render(request, "dashboard.html", {"rows": rows, "form": form, "client": client})


@login_required
def upload_tfar(request):
    """
    Upload an .xlsx TFAR file with:
      - exactly the 15 required headers, or
      - those 15 plus an optional 'client' column.
    If 'client' is present, each row must match the selected client; otherwise validation fails.
    """
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES, user=request.user)
        if not form.is_valid():
            return render(request, "upload.html", {"form": form, "error": "Invalid form submission."})

        uploaded = request.FILES.get("file")
        client_id = form.cleaned_data.get("client")

        if not uploaded:
            return render(request, "upload.html", {"form": form, "error": "Please choose a file to upload."})

        if not uploaded.name.lower().endswith(".xlsx"):
            return render(
                request,
                "upload.html",
                {"form": form, "error": "Unsupported file type. Please upload .xlsx files only."},
            )

        # Resolve client and check access
        client = get_object_or_404(Client, id=client_id)
        if not ClientMembership.objects.filter(user=request.user, client=client).exists():
            return render(request, "upload.html", {"form": form, "error": "You don't have access to this client."})

        # Load first worksheet
        try:
            wb = load_workbook(filename=uploaded, data_only=True)
            ws = wb.active
        except Exception as e:
            return render(request, "upload.html", {"form": form, "error": f"Failed to read Excel file: {e}"})

        # Read header row (row 1) — case-insensitive, spaces trimmed
        headers = [(_c.value or "").strip().lower() for _c in ws[1]]

        # Allow either 15 required OR 16 including 'client'
        def headers_ok(hdrs):
            base_ok = all(h in hdrs for h in REQUIRED_HEADERS) and len(hdrs) in (15, 16)
            return base_ok and (len(hdrs) == 15 or OPTIONAL_CLIENT_HEADER in hdrs)

        if not headers_ok(headers):
            expected = ", ".join(REQUIRED_HEADERS) + " [optional: client]"
            found = ", ".join(headers)
            return render(
                request,
                "upload.html",
                {
                    "form": form,
                    "error": f"Column mismatch.\nExpected 15 headers, or 16 including 'client'.\nRequired: {expected}\nFound: {found}",
                },
            )

        # Map header index for fast access
        idx = {name: headers.index(name) for name in headers}
        has_client_col = OPTIONAL_CLIENT_HEADER in idx

        selected_client_name = client.name.strip().lower()

        # Parse remaining rows and bulk insert
        records_to_create: list[TfarRecord] = []
        row_num = 1
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1

                # Skip empty lines (all None/empty)
                if row is None or all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                    continue

                # If present, validate client column
                if has_client_col:
                    file_client = str(row[idx[OPTIONAL_CLIENT_HEADER]] or "").strip().lower()
                    if not file_client:
                        raise ValueError("Missing client value in 'client' column")
                    if file_client != selected_client_name:
                        raise ValueError(
                            f"Client mismatch in row {row_num}: "
                            f"file has '{file_client}', selected client is '{selected_client_name}'"
                        )

                rec = TfarRecord(
                    owner=request.user,
                    client=client,
                    asset_id=_to_str(row[idx["asset id"]], 50),
                    asset_description=_to_str(row[idx["asset description"]], 250),
                    tax_start_date=_to_date(row[idx["tax start date"]]),
                    depreciation_method=_to_str(row[idx["depreciation method"]], 50),
                    purchase_cost=_to_int(row[idx["purchase cost"]]),
                    tax_effective_life=_to_int(row[idx["tax effective life"]]),
                    opening_cost=_to_int(row[idx["opening cost"]]),
                    opening_accum_depreciation=_to_int(row[idx["opening accumulated depreciation"]]),
                    opening_wdv=_to_int(row[idx["opening wdv"]]),
                    addition=_to_int(row[idx["addition"]]),
                    disposal=_to_int(row[idx["disposal"]]),
                    tax_depreciation=_to_int(row[idx["tax depreciation"]]),
                    closing_cost=_to_int(row[idx["closing cost"]]),
                    closing_accum_depreciation=_to_int(row[idx["closing accumulated depreciation"]]),
                    closing_wdv=_to_int(row[idx["closing wdv"]]),
                )
                records_to_create.append(rec)
        except ValueError as ve:
            # Provide row context for easier troubleshooting
            return render(request, "upload.html", {"form": form, "error": f"Row {row_num}: {ve}"})
        except Exception as e:
            return render(request, "upload.html", {"form": form, "error": f"Unexpected error on row {row_num}: {e}"})

        # Persist rows
        if records_to_create:
            TfarRecord.objects.bulk_create(records_to_create, batch_size=1000)

        # Remember selected client, redirect to dashboard
        request.session["selected_client_id"] = str(client.id)
        return redirect("dashboard")

    # GET
    return render(request, "upload.html", {"form": UploadForm(user=request.user)})


@login_required
def download_tfar_csv(request):
    """
    Stream the current user's TFAR rows for the selected client as CSV.
    Includes a leading 'client' column for audit/validation.
    """
    selected_client_id = request.session.get("selected_client_id")
    if not selected_client_id:
        return HttpResponse("No client selected", status=400)

    client = get_object_or_404(Client, id=selected_client_id)
    if not ClientMembership.objects.filter(user=request.user, client=client).exists():
        return HttpResponse("Forbidden", status=403)

    qs = TfarRecord.objects.filter(owner=request.user, client=client).order_by("asset_id")

    out = io.StringIO()
    out.write(",".join(["client"] + REQUIRED_HEADERS) + "\n")

    for r in qs:
        out.write(",".join([
            client.name,  # prepend client column
            r.asset_id,
            r.asset_description,
            r.tax_start_date.isoformat(),
            r.depreciation_method,
            str(r.purchase_cost),
            str(r.tax_effective_life),
            str(r.opening_cost),
            str(r.opening_accum_depreciation),
            str(r.opening_wdv),
            str(r.addition),
            str(r.disposal),
            str(r.tax_depreciation),
            str(r.closing_cost),
            str(r.closing_accum_depreciation),
            str(r.closing_wdv),
        ]) + "\n")

    response = HttpResponse(out.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{client.name}_tfar_export.csv"'
    return response


# --- Diagnostics (optional) ---
from django.conf import settings
import json

def env_view(request):
    data = {
        "ALLOWED_HOSTS": settings.ALLOWED_HOSTS,
        "CSRF_TRUSTED_ORIGINS": settings.CSRF_TRUSTED_ORIGINS,
        "DEBUG": getattr(settings, "DEBUG", None),
    }
    return HttpResponse(json.dumps(data, indent=2), content_type="application/json")


def safe_debug(request):
    try:
        rows = TfarRecord.objects.filter(owner=request.user)[:1]
        return HttpResponse("Dashboard query OK, count=" + str(rows.count()))
    except Exception as e:
        import traceback
        return HttpResponse("<pre>" + str(e) + "\n\n" + traceback.format_exc() + "</pre>", content_type="text/plain")
